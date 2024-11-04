from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message
from aiogram.types import CallbackQuery
import openpyxl
from keyboards.inlinekeyboards import get_keyboard
import pandas as pd

# Функция для записи данных в Excel
def write_to_excel(month, amount, date, status):
    # Попытка открыть существующий файл или создать новый
    try:
        workbook = openpyxl.load_workbook('/TGBot/payments.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    # Запись заголовков, если это новый файл
    if sheet.max_row == 1:
        sheet.append(['id', 'month', 'Amount', 'Date', 'status'])
    # Запись данных
    id = sheet.max_row
    sheet.append([id, month, amount, date, status])
    workbook.save('payments.xlsx')

def delete_row_to_excel(ids):
    try:
        workbook = pd.read_excel('payments.xlsx', engine='openpyxl', converters={'id':int})
        workbook = workbook[workbook['id'] != int(ids)]
        workbook.to_excel('payments.xlsx', index= False)
    except FileNotFoundError:
            print('фиаско братан')


router = Router()


class FSMadmin(StatesGroup):
     month = State()
     amount = State()
     date = State()
     status = State()
     ids = State()

@router.callback_query()
async def add_func(callback: CallbackQuery, state: FSMContext):
    if callback.data == 'add':
        await state.set_state(FSMadmin.month)
        await callback.message.reply('Напишите месяц оплаты')
    if callback.data == 'show':
        try:
             rb = openpyxl.load_workbook("payments.xlsx")
        except Exception as e:
            await callback.message.reply(f"Ошибка при загрузке файла: {e}")
            return
        sheet = rb.active
        rows = []
        for row in sheet.iter_rows(values_only=True):  # Итерируем по строкам
             # Преобразуем каждую строку в строку текста
            row_values = [str(cell) if cell is not None else '' for cell in row]  # Заменяем None на пустую строку
            rows.append(', '.join(row_values))
        await callback.message.reply('\n'.join(rows),
                                     reply_markup=get_keyboard())
    if callback.data == 'del':
        await state.set_state(FSMadmin.ids)
        await callback.message.reply('Напишите id строки для удаления')

    if callback.data == 'count':
        try:
            workbook = pd.read_excel('payments.xlsx', engine='openpyxl', converters={'id': int})
            summ = workbook['amount'].sum()
        except FileNotFoundError:
            print('фиаско братан')
        await callback.message.reply(f'За все время гараж принес: {summ} рублей')


#обработка state по добавлению
@router.message(FSMadmin.month)
async def add_date(message: Message, state: FSMContext):
     await state.update_data(month=message.text.lower())
     await state.set_state(FSMadmin.amount)
     await message.reply('Напишите сумму оплаты')

@router.message(FSMadmin.amount)
async def add_summ(message: Message, state: FSMContext):
     await state.update_data(amount=message.text.lower())
     await state.set_state(FSMadmin.date)
     await message.reply('Напишите дату оплаты')

@router.message(FSMadmin.date)
async def add_summ(message: Message, state: FSMContext):
     await state.update_data(date=message.text.lower())
     await state.set_state(FSMadmin.status)
     await message.reply('Напишите статус оплаты')

@router.message(FSMadmin.status)
async def del_func(message: Message, state: FSMContext):
     await state.update_data(status=message.text.lower())
     data = await state.get_data()
     month = data.get('month')
     amount = data.get('amount')
     date = data.get('date')
     status = data.get('status')
     try:
        write_to_excel(month, amount, date, status)
        await message.reply(f"Данные по оплате успешно внесены!",reply_markup=get_keyboard())
     except:
         await message.reply('не удалось удалить значения!', reply_markup=get_keyboard())
     await state.clear()

@router.message(FSMadmin.ids)
async def add_summ(message: Message, state: FSMContext):
     await state.update_data(ids=message.text.lower())
     data = await state.get_data()
     ids = data.get('ids')
     try:
        delete_row_to_excel(ids)
        await message.reply(f"Данные успешно удалены {ids}!", reply_markup=get_keyboard())
     except:
         await message.reply('не удалось удалить!', reply_markup=get_keyboard())
     await state.clear()




