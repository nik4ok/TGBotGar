import openpyxl

rb = openpyxl.load_workbook("C:\\TGbotGaraj\\payments.xlsx")
sheet = rb.active
print(sheet.max_row)
# rows = []
# for row in sheet.iter_rows(values_only=True):  # Итерируем по строкам
#             # Преобразуем каждую строку в строку текста
#     row_values = [str(cell) if cell is not None else '' for cell in row]  # Заменяем None на пустую строку
#     rows.append(', '.join(row_values))  # Объединяем значения в строку
#
# print(rows)
#
# # Функция для записи данных в Excel
# def write_to_excel(month, amount, date, status):
#     # Попытка открыть существующий файл или создать новый
#     try:
#         workbook = openpyxl.load_workbook('payments.xlsx')
#     except FileNotFoundError:
#         workbook = openpyxl.Workbook()
#
#     sheet = workbook.active
#     # Запись заголовков, если это новый файл
#     if sheet.max_row == 1:
#         sheet.append(['id', 'month', 'Amount', 'Date', 'status'])
#     max_id = sheet.max_row
#     # Запись данных
#     sheet.append([id, month, amount, date, status])
#     workbook.save('payments.xlsx')